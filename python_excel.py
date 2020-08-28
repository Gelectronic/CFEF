from openpyxl import Workbook
from openpyxl import load_workbook
 
wb = load_workbook(filename = "helo_world.xlsx")
   
wb.sheetnames
#['Sheet']

sheet = wb.active
sheet
#<Worksheet 'Sheet 1'>

sheet.title
#'Sheet 1'
for sheet in wb:
    print(sheet.title)

a = sheet['A1']
print(a